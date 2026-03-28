#!/usr/bin/env python3
"""
Generate an Excel spreadsheet for DeepEP shmem pool capacity estimation.

Usage:
    pip install openpyxl
    python gen_shmem_calculator.py

Output: shmem_calculator.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


def make_header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def make_input_style():
    return {
        "font": Font(bold=True, size=12, color="C00000"),
        "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "border": Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def make_result_style():
    return {
        "font": Font(bold=True, size=12, color="006100"),
        "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "border": Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def make_calc_style():
    return {
        "fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Interactive Calculator ──────────────────────────────────────────

def build_calculator_sheet(wb):
    ws = wb.active
    ws.title = "计算器"

    _set_col_widths(ws, [22, 22, 18, 50])

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11, color="4472C4")
    label_align = Alignment(horizontal="right", vertical="center")
    note_font = Font(italic=True, color="808080", size=9)

    # ── Title ──
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "DeepEP Shmem 池容量计算器"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"].value = "修改黄色单元格中的值，绿色单元格自动计算结果"
    ws["A2"].font = note_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Constants ──
    row = 4
    ws.cell(row=row, column=1, value="常量").font = section_font

    constants = [
        ("SHMEM_LOCAL_MEM_SIZE (字节)", 4 * 1024 * 1024 * 1024, "shmem 总池大小 (默认 4GB)"),
        ("SHMEM_META_DATA_SIZE (字节)", 100 * 1024 * 1024, "元数据保留 (默认 100MB)"),
    ]
    for i, (label, val, note) in enumerate(constants):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2, value=val)
        apply_style(cell, make_input_style())
        cell.number_format = "#,##0"
        ws.cell(row=r, column=4, value=note).font = note_font

    pool_cell = "B5"   # SHMEM_LOCAL_MEM_SIZE
    meta_cell = "B6"   # SHMEM_META_DATA_SIZE

    # ── Input parameters ──
    row = 8
    ws.cell(row=row, column=1, value="输入参数").font = section_font

    inputs = [
        ("hidden_size (H)", 7168, "模型隐藏维度, 如 DeepSeek-V3=7168"),
        ("num_experts (E)", 256, "专家总数, 如 DeepSeek-V3=256"),
        ("world_size (R)", 16, "EP 并行的 rank 总数"),
        ("use_quant (0或1)", 0, "0=BF16, 1=FP8/INT8 量化"),
    ]
    input_cells = {}
    for i, (label, val, note) in enumerate(inputs):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2, value=val)
        apply_style(cell, make_input_style())
        cell.number_format = "#,##0"
        ws.cell(row=r, column=4, value=note).font = note_font
        input_cells[label.split("(")[0].strip()] = f"B{r}"

    H = input_cells["hidden_size"]   # B9
    E = input_cells["num_experts"]   # B10
    R = input_cells["world_size"]    # B11
    Q = input_cells["use_quant"]     # B12

    # ── Calculation steps ──
    row = 14
    ws.cell(row=row, column=1, value="计算过程").font = section_font

    calc_style = make_calc_style()
    calc_rows = [
        ("固定开销 (字节)", f"={E}*(1+{R})*4", "E × (1 + R) × 4"),
        ("可用空间 (字节)", f"={pool_cell}-{meta_cell}-B15", "总池 - 元数据 - 固定开销"),
        ("单 token 开销 (字节)", f"={H}*2+{Q}*4", "H × 2 + (量化 ? 4 : 0)"),
        ("max_recv_tokens", f"=INT(B16/B17)", "可用空间 / 单 token 开销"),
    ]
    for i, (label, formula, note) in enumerate(calc_rows):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2)
        cell.value = formula
        cell.number_format = "#,##0"
        if i == len(calc_rows) - 1:
            apply_style(cell, make_result_style())
        else:
            apply_style(cell, calc_style)
        ws.cell(row=r, column=3, value=note).font = note_font

    # ── Tensor breakdown ──
    row = 20
    ws.cell(row=row, column=1, value="Shmem Tensor 明细").font = section_font

    tensor_header_style = make_header_style()
    tensor_headers = ["Tensor 名称", "大小 (字节)", "大小 (MB)", "说明"]
    for col, h in enumerate(tensor_headers, start=1):
        c = ws.cell(row=row + 1, column=col, value=h)
        apply_style(c, tensor_header_style)

    # T = B18 (max_recv_tokens), H = B9, E = B10, R = B11, Q = B12
    tensors = [
        ("SHMEM_META_DATA",          f"={meta_cell}",                                     "元数据保留区"),
        ("num_tokens_per_expert",     f"={E}*4",                                           "{E} × kInt"),
        ("dispatch_shmem_recv_data",  f"={R}*{E}*4",                                       "{R, E} × kInt"),
        ("combine_x (=expandx_out)",  f"=B18*{H}*2",                                      "{T, H} × BF16, expandx_out 共享此内存"),
        ("dynamic_scales_out",        f"=IF({Q}=1, B18*4, 0)",                             "{T} × kFloat, 仅量化时分配"),
    ]
    tensor_start_row = row + 2
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for i, (name, formula, desc) in enumerate(tensors):
        r = tensor_start_row + i
        ws.cell(row=r, column=1, value=name).border = thin_border
        cell_b = ws.cell(row=r, column=2)
        cell_b.value = formula
        cell_b.number_format = "#,##0"
        cell_b.border = thin_border
        cell_mb = ws.cell(row=r, column=3)
        cell_mb.value = f"=B{r}/1024/1024"
        cell_mb.number_format = "0.000"
        cell_mb.border = thin_border
        ws.cell(row=r, column=4, value=desc).border = thin_border

    # Total row
    total_row = tensor_start_row + len(tensors)
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border
    sum_range = f"B{tensor_start_row}:B{total_row - 1}"
    cell_total = ws.cell(row=total_row, column=2)
    cell_total.value = f"=SUM({sum_range})"
    cell_total.number_format = "#,##0"
    apply_style(cell_total, make_result_style())
    cell_total_mb = ws.cell(row=total_row, column=3)
    cell_total_mb.value = f"=B{total_row}/1024/1024"
    cell_total_mb.number_format = "0.000"
    apply_style(cell_total_mb, make_result_style())
    ws.cell(row=total_row, column=4, value="所有 shmem tensor 总用量").border = thin_border

    # ── Summary ──
    row = total_row + 2
    ws.cell(row=row, column=1, value="结果摘要").font = section_font

    summaries = [
        ("max_recv_tokens", "=B18", "#,##0"),
        ("可用空间 (GB)", f"=B16/1024/1024/1024", "0.000"),
        ("固定开销 (MB)", "=B15/1024/1024", "0.000"),
        ("单 token 开销 (KB)", "=B17/1024", "0.000"),
        ("实际总 shmem 用量 (GB)", f"=B{total_row}/1024/1024/1024", "0.000"),
        ("shmem 利用率", f"=B{total_row}/{pool_cell}", "0.00%"),
    ]
    for i, (label, formula, fmt) in enumerate(summaries):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2)
        cell.value = formula
        cell.number_format = fmt
        apply_style(cell, make_result_style())

    # ── Formula reference ──
    row = row + 1 + len(summaries) + 1
    ws.cell(row=row, column=1, value="公式参考").font = section_font
    formulas_text = [
        "fixed_bytes = E × (1 + R) × 4",
        "avail_bytes = SHMEM_LOCAL_MEM_SIZE - SHMEM_META_DATA_SIZE - fixed_bytes",
        "per_token_bytes = H × 2 + (use_quant ? 4 : 0)",
        "max_recv_tokens = ⌊ avail_bytes / per_token_bytes ⌋",
        "",
        "源码位置: deepep_standalone/csrc/deepep/deep_ep.cpp → Buffer::preallocate_shmem_tensors()",
        "SHMEM_LOCAL_MEM_SIZE = 4GB (硬编码常量)",
        "SHMEM_META_DATA_SIZE = 100MB (硬编码常量)",
    ]
    for i, text in enumerate(formulas_text):
        ws.cell(row=row + 1 + i, column=1, value=text).font = Font(size=10, name="Consolas")
        ws.merge_cells(start_row=row + 1 + i, start_column=1, end_row=row + 1 + i, end_column=4)


# ── Sheet 2: Reverse Calculator A — directly input max_recv_tokens ────────────

def build_reverse_direct_sheet(wb):
    ws = wb.create_sheet("反向计算器A-直接输入")

    _set_col_widths(ws, [28, 24, 18, 50])

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11, color="4472C4")
    label_align = Alignment(horizontal="right", vertical="center")
    note_font = Font(italic=True, color="808080", size=9)

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "反向计算器 A：直接输入 max_recv_tokens → 所需 shmem 池大小"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"].value = "修改黄色单元格中的值，绿色单元格自动计算所需显存池大小"
    ws["A2"].font = note_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Fixed constant ──
    row = 4
    ws.cell(row=row, column=1, value="常量").font = section_font
    ws.cell(row=5, column=1, value="SHMEM_META_DATA_SIZE (字节)").alignment = label_align
    cell = ws.cell(row=5, column=2, value=100 * 1024 * 1024)
    apply_style(cell, make_input_style())
    cell.number_format = "#,##0"
    ws.cell(row=5, column=4, value="元数据保留 (默认 100MB)").font = note_font
    meta_cell = "B5"

    # ── Input parameters ──
    row = 7
    ws.cell(row=row, column=1, value="输入参数").font = section_font
    inputs = [
        ("hidden_size (H)", 7168, "模型隐藏维度"),
        ("num_experts (E)", 256, "专家总数"),
        ("world_size (R)", 16, "EP 并行的 rank 总数"),
        ("use_quant (0或1)", 0, "0=BF16, 1=FP8/INT8 量化"),
        ("目标 max_recv_tokens", 285000, "期望支持的最大接收 token 数"),
        ("余量比例 (%)", 10, "建议申请量的额外余量百分比, 如 10 代表 10%"),
    ]
    for i, (label, val, note) in enumerate(inputs):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2, value=val)
        apply_style(cell, make_input_style())
        cell.number_format = "#,##0"
        ws.cell(row=r, column=4, value=note).font = note_font

    H = "B8"    # hidden_size
    E = "B9"    # num_experts
    R = "B10"   # world_size
    Q = "B11"   # use_quant
    T = "B12"   # target max_recv_tokens
    M = "B13"   # margin percentage

    # ── Calculation steps ──
    row = 15
    ws.cell(row=row, column=1, value="计算过程").font = section_font
    calc_style = make_calc_style()
    calc_rows = [
        ("固定开销 (字节)", f"={E}*(1+{R})*4", "E × (1 + R) × 4"),
        ("单 token 开销 (字节)", f"={H}*2+{Q}*4", "H × 2 + (量化 ? 4 : 0)"),
        ("数据区所需空间 (字节)", f"={T}*B17", "max_recv_tokens × per_token"),
        ("所需总池大小 (字节)", f"={meta_cell}+B16+B18", "元数据 + 固定开销 + 数据区"),
    ]
    for i, (label, formula, note) in enumerate(calc_rows):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2)
        cell.value = formula
        cell.number_format = "#,##0"
        if i == len(calc_rows) - 1:
            apply_style(cell, make_result_style())
        else:
            apply_style(cell, calc_style)
        ws.cell(row=r, column=3, value=note).font = note_font

    # ── Tensor breakdown ──
    row = 21
    ws.cell(row=row, column=1, value="Shmem Tensor 明细").font = section_font
    tensor_headers = ["Tensor 名称", "大小 (字节)", "大小 (MB)", "说明"]
    for col, h_text in enumerate(tensor_headers, start=1):
        c = ws.cell(row=row + 1, column=col, value=h_text)
        apply_style(c, make_header_style())

    tensors = [
        ("SHMEM_META_DATA",          f"={meta_cell}",                   "元数据保留区"),
        ("num_tokens_per_expert",     f"={E}*4",                         "{E} × kInt"),
        ("dispatch_shmem_recv_data",  f"={R}*{E}*4",                     "{R, E} × kInt"),
        ("combine_x (=expandx_out)",  f"={T}*{H}*2",                    "{T, H} × BF16, expandx_out 共享此内存"),
        ("dynamic_scales_out",        f"=IF({Q}=1, {T}*4, 0)",          "{T} × kFloat, 仅量化时分配"),
    ]
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    tensor_start_row = row + 2
    for i, (name, formula, desc) in enumerate(tensors):
        r = tensor_start_row + i
        ws.cell(row=r, column=1, value=name).border = thin_border
        cell_b = ws.cell(row=r, column=2)
        cell_b.value = formula
        cell_b.number_format = "#,##0"
        cell_b.border = thin_border
        cell_mb = ws.cell(row=r, column=3)
        cell_mb.value = f"=B{r}/1024/1024"
        cell_mb.number_format = "0.000"
        cell_mb.border = thin_border
        ws.cell(row=r, column=4, value=desc).border = thin_border

    total_row = tensor_start_row + len(tensors)
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border
    sum_range = f"B{tensor_start_row}:B{total_row - 1}"
    cell_total = ws.cell(row=total_row, column=2)
    cell_total.value = f"=SUM({sum_range})"
    cell_total.number_format = "#,##0"
    apply_style(cell_total, make_result_style())
    cell_total_mb = ws.cell(row=total_row, column=3)
    cell_total_mb.value = f"=B{total_row}/1024/1024"
    cell_total_mb.number_format = "0.000"
    apply_style(cell_total_mb, make_result_style())
    ws.cell(row=total_row, column=4, value="所有 shmem tensor 总用量").border = thin_border

    # ── Result summary ──
    row = total_row + 2
    ws.cell(row=row, column=1, value="结果摘要").font = section_font
    summaries = [
        ("所需 shmem 池大小 (GB)", "=B19/1024/1024/1024", "0.000"),
        ("所需 shmem 池大小 (MB)", "=B19/1024/1024", "0.000"),
        ("建议申请量 (GB, 含余量)", f"=CEILING(B19*(1+{M}/100)/1024/1024/1024, 0.5)", "0.0"),
        ("固定开销占比", f"=B16/B19", "0.000%"),
        ("元数据占比", f"={meta_cell}/B19", "0.00%"),
        ("数据区占比", "=B18/B19", "0.00%"),
    ]
    for i, (label, formula, fmt) in enumerate(summaries):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2)
        cell.value = formula
        cell.number_format = fmt
        apply_style(cell, make_result_style())

    # ── Formula reference ──
    row = row + 1 + len(summaries) + 1
    ws.cell(row=row, column=1, value="公式参考").font = section_font
    formulas_text = [
        "required_pool = META_SIZE + E×(1+R)×4 + max_recv_tokens × (H×2 + quant×4)",
        "",
        "反向推导自:",
        "max_recv_tokens = ⌊ (pool - META_SIZE - E×(1+R)×4) / (H×2 + quant×4) ⌋",
    ]
    for i, text in enumerate(formulas_text):
        ws.cell(row=row + 1 + i, column=1, value=text).font = Font(size=10, name="Consolas")
        ws.merge_cells(start_row=row + 1 + i, start_column=1, end_row=row + 1 + i, end_column=4)


# ── Sheet 3: Reverse Calculator B — estimate from seqLen/topk/imbalance ──────

def build_reverse_estimate_sheet(wb):
    ws = wb.create_sheet("反向计算器B-推算")

    _set_col_widths(ws, [28, 24, 18, 50])

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11, color="4472C4")
    label_align = Alignment(horizontal="right", vertical="center")
    note_font = Font(italic=True, color="808080", size=9)

    # ── Title ──
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "反向计算器 B：由 seqLen/topk/不均匀比例 推算所需 shmem 池大小"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"].value = "修改黄色单元格中的值，绿色单元格自动计算所需显存池大小"
    ws["A2"].font = note_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Fixed constant ──
    row = 4
    ws.cell(row=row, column=1, value="常量").font = section_font

    ws.cell(row=5, column=1, value="SHMEM_META_DATA_SIZE (字节)").alignment = label_align
    cell = ws.cell(row=5, column=2, value=100 * 1024 * 1024)
    apply_style(cell, make_input_style())
    cell.number_format = "#,##0"
    ws.cell(row=5, column=4, value="元数据保留 (默认 100MB)").font = note_font
    meta_cell = "B5"

    # ── Input parameters ──
    row = 7
    ws.cell(row=row, column=1, value="输入参数").font = section_font

    inputs = [
        ("hidden_size (H)", 7168, "模型隐藏维度"),
        ("num_experts (E)", 256, "专家总数"),
        ("world_size (R)", 16, "EP 并行的 rank 总数"),
        ("use_quant (0或1)", 0, "0=BF16, 1=FP8/INT8 量化"),
        ("seqLen (每卡输入token数)", 4096, "dispatch 输入的序列长度 (每张卡)"),
        ("topk", 8, "每个 token 路由到的专家数"),
        ("不均匀比例", 1.5, "负载不均匀倍数, 1.0=完全均匀, 1.5=最坏比均值多50%"),
    ]
    for i, (label, val, note) in enumerate(inputs):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2, value=val)
        apply_style(cell, make_input_style())
        if isinstance(val, float):
            cell.number_format = "0.0"
        else:
            cell.number_format = "#,##0"
        ws.cell(row=r, column=4, value=note).font = note_font

    H = "B8"      # hidden_size
    E = "B9"      # num_experts
    R = "B10"     # world_size
    Q = "B11"     # use_quant
    SL = "B12"    # seqLen
    TK = "B13"    # topk
    IM = "B14"    # imbalance ratio

    # ── Calculation steps ──
    row = 16
    ws.cell(row=row, column=1, value="计算过程").font = section_font

    calc_style = make_calc_style()
    # max_recv_tokens = seqLen × topk × 不均匀比例
    # (每卡发出 seqLen*topk 个token-expert对, 均匀分配时每卡收到 seqLen*topk,
    #  乘以不均匀比例得到最坏情况接收量)
    calc_rows = [
        ("max_recv_tokens (估算)", f"=CEILING({SL}*{TK}*{IM}, 1)", "seqLen × topk × 不均匀比例"),
        ("固定开销 (字节)", f"={E}*(1+{R})*4", "E × (1 + R) × 4"),
        ("单 token 开销 (字节)", f"={H}*2+{Q}*4", "H × 2 + (量化 ? 4 : 0)"),
        ("数据区所需空间 (字节)", f"=B17*B19", "max_recv_tokens × per_token"),
        ("所需总池大小 (字节)", f"={meta_cell}+B18+B20", "元数据 + 固定开销 + 数据区"),
    ]
    for i, (label, formula, note) in enumerate(calc_rows):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2)
        cell.value = formula
        cell.number_format = "#,##0"
        if i == 0:
            apply_style(cell, make_result_style())
        elif i == len(calc_rows) - 1:
            apply_style(cell, make_result_style())
        else:
            apply_style(cell, calc_style)
        ws.cell(row=r, column=3, value=note).font = note_font

    # ── Tensor breakdown ──
    # T_cell = B17 (max_recv_tokens)
    T_cell = "B17"
    row = 23
    ws.cell(row=row, column=1, value="Shmem Tensor 明细").font = section_font

    tensor_header_style = make_header_style()
    tensor_headers = ["Tensor 名称", "大小 (字节)", "大小 (MB)", "说明"]
    for col, h_text in enumerate(tensor_headers, start=1):
        c = ws.cell(row=row + 1, column=col, value=h_text)
        apply_style(c, tensor_header_style)

    tensors = [
        ("SHMEM_META_DATA",          f"={meta_cell}",                                     "元数据保留区"),
        ("num_tokens_per_expert",     f"={E}*4",                                           "{E} × kInt"),
        ("dispatch_shmem_recv_data",  f"={R}*{E}*4",                                       "{R, E} × kInt"),
        ("combine_x (=expandx_out)",  f"={T_cell}*{H}*2",                                  "{T, H} × BF16, expandx_out 共享此内存"),
        ("dynamic_scales_out",        f"=IF({Q}=1, {T_cell}*4, 0)",                         "{T} × kFloat, 仅量化时分配"),
    ]
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    tensor_start_row = row + 2
    for i, (name, formula, desc) in enumerate(tensors):
        r = tensor_start_row + i
        ws.cell(row=r, column=1, value=name).border = thin_border
        cell_b = ws.cell(row=r, column=2)
        cell_b.value = formula
        cell_b.number_format = "#,##0"
        cell_b.border = thin_border
        cell_mb = ws.cell(row=r, column=3)
        cell_mb.value = f"=B{r}/1024/1024"
        cell_mb.number_format = "0.000"
        cell_mb.border = thin_border
        ws.cell(row=r, column=4, value=desc).border = thin_border

    total_row = tensor_start_row + len(tensors)
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border
    sum_range = f"B{tensor_start_row}:B{total_row - 1}"
    cell_total = ws.cell(row=total_row, column=2)
    cell_total.value = f"=SUM({sum_range})"
    cell_total.number_format = "#,##0"
    apply_style(cell_total, make_result_style())
    cell_total_mb = ws.cell(row=total_row, column=3)
    cell_total_mb.value = f"=B{total_row}/1024/1024"
    cell_total_mb.number_format = "0.000"
    apply_style(cell_total_mb, make_result_style())
    ws.cell(row=total_row, column=4, value="所有 shmem tensor 总用量").border = thin_border

    # ── Result summary ──
    row = total_row + 2
    ws.cell(row=row, column=1, value="结果摘要").font = section_font

    # B21 = 所需总池大小, B17 = max_recv_tokens
    summaries = [
        ("估算 max_recv_tokens", "=B17", "#,##0"),
        ("所需 shmem 池大小 (GB)", "=B21/1024/1024/1024", "0.000"),
        ("所需 shmem 池大小 (MB)", "=B21/1024/1024", "0.000"),
        ("建议申请量 (GB, 向上取整)", "=CEILING(B21/1024/1024/1024, 0.5)", "0.0"),
        ("固定开销占比", f"=B18/B21", "0.000%"),
        ("元数据占比", f"={meta_cell}/B21", "0.00%"),
        ("数据区占比", "=B20/B21", "0.00%"),
    ]
    for i, (label, formula, fmt) in enumerate(summaries):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).alignment = label_align
        cell = ws.cell(row=r, column=2)
        cell.value = formula
        cell.number_format = fmt
        apply_style(cell, make_result_style())

    # ── Formula reference ──
    row = row + 1 + len(summaries) + 1
    ws.cell(row=row, column=1, value="公式参考").font = section_font
    formulas_text = [
        "max_recv_tokens = ⌈ seqLen × topk × 不均匀比例 ⌉",
        "required_pool = META_SIZE + E×(1+R)×4 + max_recv_tokens × (H×2 + quant×4)",
        "",
        "不均匀比例说明:",
        "  1.0 = 完全均匀, 每卡收到 seqLen×topk 个 token",
        "  1.5 = 最坏情况比均值多 50%",
        "  2.0 = 最坏情况是均值的 2 倍",
    ]
    for i, text in enumerate(formulas_text):
        ws.cell(row=row + 1 + i, column=1, value=text).font = Font(size=10, name="Consolas")
        ws.merge_cells(start_row=row + 1 + i, start_column=1, end_row=row + 1 + i, end_column=4)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    build_calculator_sheet(wb)
    build_reverse_direct_sheet(wb)
    build_reverse_estimate_sheet(wb)

    out = "shmem_calculator.xlsx"
    wb.save(out)
    print(f"✓ 已生成: {out}")
    print("  Sheet 1 [正向计算器]         — 输入 H/E/R/量化 → 算 max_recv_tokens + Tensor 明细")
    print("  Sheet 2 [反向计算器A-直接输入] — 直接输入 max_recv_tokens → 算所需池大小")
    print("  Sheet 3 [反向计算器B-推算]     — 输入 seqLen/topk/不均匀比例 → 推算所需池大小")


if __name__ == "__main__":
    main()
